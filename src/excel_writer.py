"""Write PtoEntry records to a month-per-tab calendar workbook.

One sheet per month in the date range (June, July, ...). Each sheet is a
Mon..Sun calendar grid: for every week, a date-label row followed by a names
row listing everyone on PTO that day (one name per line). Days outside the
sheet's month are shown greyed. A 'Notes' row sits at the bottom.
"""
from __future__ import annotations

import calendar as _calendar
from collections import defaultdict
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import PtoEntry

WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DATE_FILL = PatternFill("solid", fgColor="DDEBF7")
DATE_FONT = Font(bold=True, color="1F1F1F")
SPILL_FONT = Font(bold=True, color="A6A6A6")   # greyed out-of-month dates
NAME_FONT = Font(color="1F1F1F")
WEEKEND_FILL = PatternFill("solid", fgColor="F2F2F2")
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COL_WIDTH = 18
LINE_HEIGHT = 15
MIN_NAME_ROW_HEIGHT = 72


def write_report(
    entries: list[PtoEntry],
    out_path: str,
    start: datetime,
    end: datetime,
) -> str:
    by_date = _index_by_date(entries)
    multi_year = start.year != end.year

    wb = Workbook()
    wb.remove(wb.active)
    for year, month in _months_in_range(start, end):
        title = f"{_calendar.month_name[month]} {year}" if multi_year else _calendar.month_name[month]
        _write_month(wb.create_sheet(title[:31]), year, month, by_date)
    if not wb.sheetnames:
        wb.create_sheet("No PTO")
    wb.save(out_path)
    return out_path


def _index_by_date(entries: list[PtoEntry]) -> dict[date, list[tuple[str, bool]]]:
    """date -> sorted list of (person, tentative)."""
    out: dict[date, list[tuple[str, bool]]] = defaultdict(list)
    for e in entries:
        for d in e.dates:
            out[d].append((e.person, e.tentative))
    for d in out:
        out[d].sort(key=lambda t: t[0].lower())
    return out


def _write_month(ws, year: int, month: int, by_date: dict) -> None:
    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

    # Weekday header row.
    for c, wd in enumerate(WEEKDAYS, start=1):
        cell = ws.cell(1, c, wd)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    ws.freeze_panes = "A2"

    weeks = _calendar.Calendar(firstweekday=0).monthdatescalendar(year, month)
    row = 2
    for week in weeks:
        max_names = 1
        for c, d in enumerate(week, start=1):
            in_month = d.month == month
            weekend = d.weekday() >= 5

            # date-label cell
            lab = ws.cell(row, c, d.strftime("%d-%b"))
            lab.font = DATE_FONT if in_month else SPILL_FONT
            lab.fill = DATE_FILL
            lab.alignment = Alignment(horizontal="left")
            lab.border = BORDER

            # names cell (only for days belonging to this month)
            names = []
            if in_month:
                names = [f"{p}(tentative)" if tentative else p
                         for p, tentative in by_date.get(d, [])]
            nm = ws.cell(row + 1, c, "\n".join(names))
            nm.font = NAME_FONT
            nm.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            nm.border = BORDER
            if weekend:
                nm.fill = WEEKEND_FILL
            max_names = max(max_names, len(names))

        ws.row_dimensions[row + 1].height = max(MIN_NAME_ROW_HEIGHT, LINE_HEIGHT * max_names + 6)
        row += 2

    ws.cell(row + 1, 1, "Notes").font = Font(bold=True)
    ws.sheet_view.showGridLines = False


def _months_in_range(start: datetime, end: datetime):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield y, m
        m += 1
        if m > 12:
            m = 1
            y += 1
